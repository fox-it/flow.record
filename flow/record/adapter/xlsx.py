from __future__ import annotations

from base64 import b64decode, b64encode
from datetime import datetime, timezone
from typing import TYPE_CHECKING, Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

from flow import record
from flow.record import fieldtypes
from flow.record.adapter import AbstractReader, AbstractWriter
from flow.record.fieldtypes.net import ipaddress
from flow.record.selector import make_selector
from flow.record.utils import is_stdout

if TYPE_CHECKING:
    from collections.abc import Iterator
    from pathlib import Path

    from flow.record.base import Record

__usage__ = """
Microsoft Excel spreadsheet adapter
---
Write usage: rdump -w xlsx://[PATH]
Read usage: rdump xlsx://[PATH]
[PATH]: path to file. Leave empty or "-" to output to stdout
"""


def sanitize_fieldvalues(values: Iterator[Any]) -> Iterator[Any]:
    """Sanitize field values so openpyxl will accept them."""

    for value in values:
        # openpyxl doesn't support timezone-aware datetime instances,
        # so we convert to UTC and then remove the timezone info.
        if isinstance(value, datetime) and value.tzinfo is not None:
            value = value.astimezone(timezone.utc).replace(tzinfo=None)

        elif type(value) in [ipaddress, list, fieldtypes.posix_path, fieldtypes.windows_path]:
            value = str(value)

        elif isinstance(value, bytes):
            base64_encode = False
            try:
                new_value = 'b"' + value.decode(errors="surrogateescape") + '"'
                if ILLEGAL_CHARACTERS_RE.search(new_value):
                    base64_encode = True
                else:
                    value = new_value
            except UnicodeDecodeError:
                base64_encode = True
            if base64_encode:
                value = "base64:" + b64encode(value).decode()

        yield value


class XlsxWriter(AbstractWriter):
    fp = None
    wb = None

    def __init__(self, path: str | Path | BinaryIO, **kwargs):
        self.fp = record.open_path_or_stream(path, "wb")
        self.wb = Workbook()
        self.ws = self.wb.active

        # Remove the active work sheet, every Record Descriptor will have its own sheet.
        self.wb.remove(self.ws)
        self.descs = []
        self._last_dec = None

    def write(self, r: Record) -> None:
        if r._desc not in self.descs:
            self.descs.append(r._desc)
            ws = self.wb.create_sheet(r._desc.name.strip().replace("/", "-"))
            field_types = []
            field_names = []

            for field_name, field in r._desc.get_all_fields().items():
                field_types.append(field.typename)
                field_names.append(field_name)

            ws.append(field_types)
            ws.append(field_names)

        if r._desc != self._last_dec:
            self._last_dec = r._desc
            self.ws = self.wb[r._desc.name.strip().replace("/", "-")]

        values = list(sanitize_fieldvalues(value for value in r._asdict().values()))

        try:
            self.ws.append(values)
        except ValueError as e:
            raise ValueError(f"Unable to write values to workbook: {e!s}")

    def flush(self) -> None:
        if self.wb:
            self.wb.save(self.fp)

    def close(self) -> None:
        if self.wb:
            self.wb.close()
        self.wb = None

        if self.fp and not is_stdout(self.fp):
            self.fp.close()
        self.fp = None


class XlsxReader(AbstractReader):
    fp = None

    def __init__(self, path: str | Path | BinaryIO, selector: str | None = None, **kwargs):
        self.selector = make_selector(selector)
        self.fp = record.open_path_or_stream(path, "rb")
        self.desc = None
        self.wb = load_workbook(self.fp)
        self.ws = self.wb.active

    def close(self) -> None:
        if self.fp:
            self.fp.close()
        self.fp = None

    def __iter__(self) -> Iterator[Record]:
        for worksheet in self.wb.worksheets:
            desc = None
            desc_name = worksheet.title.replace("-", "/")
            field_names = None
            field_types = None
            for row in worksheet:
                if field_types is None:
                    field_types = [col.value for col in row if col.value]
                    continue
                if field_names is None:
                    field_names = [
                        col.value.replace(" ", "_").lower()
                        for col in row
                        if col.value and not col.value.startswith("_")
                    ]
                    desc = record.RecordDescriptor(desc_name, list(zip(field_types, field_names)))
                    continue

                record_values = []
                for idx, col in enumerate(row):
                    value = col.value
                    if field_types[idx] == "bytes":
                        if value[1] == '"':  # If so, we know this is b""
                            # Cut of the b" at the start and the trailing "
                            value = value[2:-1].encode(errors="surrogateescape")
                        else:
                            # If not, we know it is base64 encoded (so we cut of the starting 'base64:')
                            value = b64decode(value[7:])
                    record_values.append(value)
                obj = desc(*record_values)
                if not self.selector or self.selector.match(obj):
                    yield obj
